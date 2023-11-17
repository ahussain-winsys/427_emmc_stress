import openpyxl
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt

def create_line_graph(x_values, y_values, excel_file_path, image_path='line_graph.png'):
    # Create a line graph using matplotlib
    plt.plot(x_values, y_values)
    plt.xlabel('X-axis')
    plt.ylabel('Y-axis')
    plt.title('Sample Line Graph')
    plt.grid(True)

    # Save the matplotlib figure to an image file
    plt.savefig(image_path)
    plt.close()  # Close the matplotlib figure to release resources

    # Create an Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add data to the worksheet
    for row, value in enumerate(x_values, start=1):
        worksheet.cell(row=row, column=1, value=value)

    for row, value in enumerate(y_values, start=1):
        worksheet.cell(row=row, column=2, value=value)

    # Add a chart to the worksheet
    chart = LineChart()
    chart.add_data(Reference(worksheet, min_col=2, min_row=1, max_col=2, max_row=len(y_values)))
    chart.set_categories(Reference(worksheet, min_col=1, min_row=2, max_row=len(x_values)))

    # Add the chart to the worksheet
    worksheet.add_chart(chart, "D2")

    # Save the workbook to an Excel file
    workbook.save(excel_file_path)

    print(f"Excel file with graph saved at: {excel_file_path}")